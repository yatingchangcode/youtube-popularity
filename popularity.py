# import libraries 
from apiclient.discovery import build 
import pprint 
import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
from urllib.parse import urlparse
from urllib.parse import parse_qs
from datetime import date
from datetime import datetime
from openpyxl.utils import get_column_letter
import os
import tkinter as tk
from tkinter import messagebox

# arguments to be passed to build function 
API_KEY = "AIzaSyAFClD5KHqBZoKQmk6hhFUuuN9kEAkIAHY"
YOUTUBE_API_SERVICE_NAME = "youtube"
YOUTUBE_API_VERSION = "v3"

# creating youtube resource 
# object for interacting with API 
youtube = build(YOUTUBE_API_SERVICE_NAME, 
        YOUTUBE_API_VERSION, 
        developerKey = API_KEY) 

def check_dateidx(sheet):
    check = 1
    row_idx = 2
    while (check):
        if not sheet.cell(row=row_idx, column=1).value: #which is null
            check = 0
            return row_idx
        row_idx += 1

def initWorkbook(sheet):
    sheet.column_dimensions[get_column_letter(1)].width = 10.0
    sheet.column_dimensions[get_column_letter(2)].width = 30.0
    sheet.column_dimensions[get_column_letter(3)].width = 20.0
    sheet.column_dimensions[get_column_letter(4)].width = 10.0
    sheet.column_dimensions[get_column_letter(5)].width = 10.0
    sheet.column_dimensions[get_column_letter(6)].width = 10.0
    sheet.column_dimensions[get_column_letter(7)].width = 10.0
    sheet.column_dimensions[get_column_letter(8)].width = 10.0
    sheet.column_dimensions[get_column_letter(9)].width = 10.0
    sheet.column_dimensions[get_column_letter(10)].width = 10.0
    sheet.column_dimensions[get_column_letter(11)].width = 30.0
    sheet.column_dimensions[get_column_letter(12)].width = 30.0

    sheet.cell(row=1, column=1).value = "DATE"
    sheet.cell(row=1, column=2).value = "TITLE"
    sheet.cell(row=1, column=3).value = "URL"
    sheet.cell(row=1, column=4).value = "viewCount"
    sheet.cell(row=1, column=5).value = "likeCount"
    sheet.cell(row=1, column=6).value = "dislikeCount"
    sheet.cell(row=1, column=7).value = "favoriteCount"
    sheet.cell(row=1, column=8).value = "commentCount"
    sheet.cell(row=1, column=9).value = "subscriberCount"
    sheet.cell(row=1, column=10).value = "tag"
    sheet.cell(row=1, column=11).value = "description"
    sheet.cell(row=1, column=12).value = "contentDetails"

def setDataToExcel(sheet, cur_date, dateidx, result, channel_rlt):
    sheet.cell(row=dateidx, column=1).value = cur_date
    sheet.cell(row=dateidx, column=2).value = result["snippet"]["title"]
    sheet.cell(row=dateidx, column=3).value = "https://www.youtube.com/watch?v=" + result["id"]
    sheet.cell(row=dateidx, column=4).value = result["statistics"]["viewCount"]
    sheet.cell(row=dateidx, column=5).value = result["statistics"]["likeCount"]
    sheet.cell(row=dateidx, column=6).value = result["statistics"]["dislikeCount"]
    sheet.cell(row=dateidx, column=7).value = result["statistics"]["favoriteCount"]
    sheet.cell(row=dateidx, column=8).value = result["statistics"]["commentCount"]
    sheet.cell(row=dateidx, column=9).value = channel_rlt["statistics"]["subscriberCount"]
    sheet.cell(row=dateidx, column=10).value = str(result["snippet"]["tags"])
    sheet.cell(row=dateidx, column=11).value = result['snippet']['description'] 
    sheet.cell(row=dateidx, column=12).value = str(result['contentDetails']) 

def multiple_video_details(sheet, urlIDs): 

    dateidx = check_dateidx(sheet)
    print("dateidx %s" % dateidx)

    today = date.today()
    cur_date = today.strftime("%y/%m/%d")
    # Call the videos.list method
    # to retrieve video info
    column_idx = 2

    for ID in urlIDs:
        list_videos_byid = youtube.videos().list( 
                    id = ID, 
                    part = "id, snippet, contentDetails, statistics", 
	).execute()

    # extracting the results from search response 
        results = list_videos_byid.get("items", []) 
        # empty list to store video details 
        videos = []

        for result in results:
            column = get_column_letter(column_idx)
            #sheet.column_dimensions[column].width = 40.0
            result_str = ""
            result_statistics = result["statistics"]
            for key in result_statistics:
                result_str += key + ' : ' + result_statistics[key] + '\n'

            #search channel subscribercount
            channel = youtube.channels().list(
                    part="statistics",
                    id= result["snippet"]["channelId"]
            ).execute()

            channel_rltss = channel.get("items", [])
            for channel_rlt in channel_rltss:
                result_str += 'subscriberCount' + ' : ' + channel_rlt["statistics"]["subscriberCount"]
            #======================
            setDataToExcel(sheet, cur_date, dateidx, result, channel_rlt)
            dateidx += 1
            videos.append("TITLE:(% s)\n TAG:(% s)\n PUBLICDATE:(% s)\n CONTENT:(% s)\n STAT:(% s)\n\n\n"
        						% (result["snippet"]["title"], 
        								result["snippet"]["tags"], 
        								#result['snippet']['description'], 
        								result["snippet"]["publishedAt"], 
        								result['contentDetails'], 
        								result["statistics"])) 
            column_idx = column_idx + 1
        	
        print ("Videos:\n", "\n".join(videos), "\n")

# created youtube id parser without regexp:
# This one is great for parsing all of the possible youtube link formats.
def url_to_video_id(value):
    query = urlparse(value)
    if query.hostname == 'youtu.be':
        return query.path[1:]
    if query.hostname in ('www.youtube.com', 'youtube.com'):
        if query.path == '/watch':
            p = parse_qs(query.query)
            return p['v'][0]
        if query.path[:7] == '/embed/':
            return query.path.split('/')[2]
        if query.path[:3] == '/v/':
            return query.path.split('/')[2]
    # fail?
    return None

def parse_videoID(sheet, url_idx):
    videoIDs = []
    max_row = sheet.max_row
    for i in range(2, max_row + 1):
        url = sheet.cell(row=i, column=url_idx).value
        if url:
            videoID = url_to_video_id(url)
            videoIDs.append(videoID)
    return videoIDs

def read_xlxs(excel_file):
    wb = load_workbook(excel_file)
    sheet = wb.active
    max_column = sheet.max_column
    for i in range(1, max_column + 1):
    	if sheet.cell(row=1, column=i).value == "url" :
    		url_idx = i
    print ('url_idx =', url_idx)
    print('\n')
    return parse_videoID(sheet, url_idx)

def main():
    file_name = "popularity.xlsx"
    urlIDs = read_xlxs("data.xlsx")
    newfile = 1

    if os.path.exists(file_name):
        newfile = 0
        wb = load_workbook(file_name)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        initWorkbook(sheet)

    multiple_video_details(sheet, urlIDs)
    wb.save(filename = file_name)

if __name__ == "__main__":

    app_date = datetime(year=2020,month=3,day=16) #setup a datetime object
    now = datetime.now()
    if (now-app_date).days >=5: #change to 30
        messagebox.showerror("Error","Your tool had expired")
    else:
        main()


#multiple_video_details() 
